import openpyxl


class GetTestsData:
    @staticmethod
    def get_tests_data(data_file_name, worksheet_name, scenario_name):
        result_dict = {}
        workbook = openpyxl.load_workbook("./data/"+data_file_name+".xlsx")
        worksheet = workbook[worksheet_name]
        for i in range (1, worksheet.max_row + 1):  # To get rows
            if worksheet.cell(row=i, column=2).value == scenario_name:
                for j in range(1, worksheet.max_column + 1):  # To get columns
                    result_dict[worksheet.cell(row=1, column=j).value] = worksheet.cell(row=i, column=j).value
        return [result_dict]
