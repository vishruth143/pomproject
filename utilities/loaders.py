from xlsxwriter import Workbook


def load_json(path: str):
    import json

    with open(path) as f:
        return json.load(f)


def load_yaml(path: str):
    import yaml

    with open(path) as f:
        return yaml.load(f, Loader=yaml.FullLoader)


def load_csv(path: str):
    import csv

    with open(path) as f:
        header_row, *data_rows = csv.reader(f)

        rows = []

        for data_row in data_rows:
            data = dict(zip(header_row, data_row))
            rows.append(data)

        return rows


def configure_xlsx_loader(*args, **kwargs):
    """
    Factory function to generate configured xlsx loader function

    :param config:
    :return: Configured xlsx loader function
    """

    data_only = kwargs.get("data_only", True)
    strip_header_cells = kwargs.get("strip_header_cells", True)
    strip_data_cells = kwargs.get("strip_data_cells", True)

    def configure_xlsx_loader(path: str):
        import openpyxl

        data = {}

        # This file is generated while the file is open in Excel
        # and should be ignored as its not a valid spreadsheet
        if "~$" in path:
            return None

        wb = openpyxl.load_workbook(path, read_only=True, data_only=data_only)

        for sheet in wb.sheetnames:
            data[sheet] = []
            raw_header_row, *raw_data_rows = wb[sheet].rows

            header_names = []

            for cell in raw_header_row:
                cell_value = cell.value
                if strip_header_cells:
                    if isinstance(cell_value, str):
                        cell_value = cell_value.strip()
                    header_names.append(cell_value)

            for raw_data_row in raw_data_rows:
                cell_values = []

                for cell in raw_data_row:
                    cell_value = cell.value
                    if strip_data_cells:
                        if isinstance(cell_value, str):
                            cell_value = cell_value.strip()
                    cell_values.append(cell_value)

                data_row = dict(zip(header_names, cell_values))
                data[sheet].append(data_row)

        return data

    return configure_xlsx_loader


def load_xlsx(path: str):
    """
    Load an xlsx file from a path
    :param path:
    :return: Dictionary of data from sheets with sheet name as keys
    """

    return configure_xlsx_loader()(path)


def write_data_dic_to_file(data_dic):
    output_data_file_path = 'output/nopcommerce_test_data_output.xlsx'
    wb = Workbook(output_data_file_path)

    # Excel header formatting
    header_format = wb.add_format({'bold': True, 'font_color': 'white', 'bg_color': 'balck'})

    for sheet in data_dic:
        ordered_list = list(data_dic[sheet][0].keys())

        ws = wb.add_worksheet(sheet)

        first_row = 0
        col = 0
        for header in ordered_list:
            col = ordered_list.index(header)  # Keeping the order
            ws.write(first_row, col, header, header_format)  # Write first row which is the header of the worksheet

        row = 1
        for data_row in data_dic[sheet]:
            for _key, _value in data_row.items():
                col = ordered_list.index(_key)
                ws.write(row, col, _value)
            row += 1  # Enter the next row
    wb.close()
